# windows10 python3.12 ortools9.10 不能运行，原因未知。
from ortools.sat.python import cp_model
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font, PatternFill, Alignment,Border,Side
#from dateutil.parser import parse
#import calendar
import os
current_dir=os.path.dirname(os.path.abspath(__file__))
model=cp_model.CpModel()

# read data from excel(文件名)
def read_data_from_excel(filename):
    workbook=load_workbook(filename,data_only=True)
    duty_count=0
    rest_count=0
    lastweek_assignments=[]
    for row in workbook['上月班次衔接'].iter_rows(min_row=2,min_col=2,max_col=4,values_only=True):
        lastweek_assignments.append(tuple(row))
    
    employees=[]
    for row in workbook['capital'].iter_rows(min_row=2,min_col=1,max_col=1,values_only=True):
        employees.append(row[0])
    work_days=workbook['capital']['C2'].value
    num_days=workbook['capital']['F2'].value
    
    #date=parse(date)
    #num_days=calendar.monthrange(date.year,date.month)
    holi_days=[]
    for row in workbook['capital'].iter_rows(min_row=2,max_row=6,min_col=7,max_col=7,values_only=True):
        if row[0] is not None:
            holi_days.append(row[0])
    min_M910=workbook['capital']['P2'].value
    min_night=workbook['capital']['Q2'].value

    dayly_cover_demands=[]
    for row in workbook['班次要求'].iter_rows(min_row=2,min_col=1,max_col=10,values_only=True):
        if row[0] is not None:
            dayly_cover_demands.append(tuple(row))
    
    requests=[]
    for row in workbook['班次意见'].iter_rows(min_row=2,min_col=2,max_col=5,values_only=True):
        # 因为sheet中已用行数可能大于requests行数，所以要判断。班次要求同理。
        if row[0] is not None:
            requests.append(tuple(row))
            if row[2]==0:
                rest_count+=1
            else:
                duty_count+=1
    
    fixed_assignments=[]
    for row in workbook['固定班次'].iter_rows(min_row=2,min_col=2,max_col=4,values_only=True):
        fixed_assignments.append(tuple(row))

    workbook.close()
    return lastweek_assignments, employees, work_days, num_days, holi_days, min_M910, min_night, dayly_cover_demands, requests, fixed_assignments,rest_count,duty_count

# 读入各项参数
filename='mvtm_zx_nohalf.xlsm'
file_path=os.path.join(current_dir,filename)
lastweek_assignments, employees, work_days, num_days, holi_days, min_M910, min_night, dayly_cover_demands, requests, fixed_assignments,rest_count,duty_count=read_data_from_excel(file_path)
print("requests readed")
print(f"-rest  :{rest_count}")
print(f"-duty  :{duty_count}")
num_employees=len(employees)
shifts=['休','M1','M2','M3','M4','VM1','VM2','VM3','M8','M9','M10','年假','培训']
num_shifts=len(shifts)

obj_bool_vars=[]
obj_bool_coeffs=[]
# 定义排班变量
work={}
for e in range(num_employees):
    for s in range(num_shifts):
        for d in range(-6,num_days):
            work[e,d,s]=model.NewBoolVar('work%i_%i_%i' % (e,d,s))


# fixed_assignments(e,s,d)
px=shifts.index("培训")
for e,d,s in fixed_assignments:
#    if s==px:
#        work[e,d,s]=model.NewBoolVar('work%i_%i_%i' % (e,d,s))
    model.Add(work[e,d,s]==1)
# lastweek_assignments=[]
for e,d,s in lastweek_assignments:
    model.Add(work[e,d,s]==1)
# requests(e,s,d,w)
for e,d,s,w in requests:
    obj_bool_vars.append(work[e,d,s])
    obj_bool_coeffs.append(w)

# 每个员工每天只能上一个班次
for e in range(num_employees):
    for d in range(num_days):
        model.Add(sum(work[e,d,s] for s in range(num_shifts))==1)

# 最大连续工作天数不超过7天
for e in range(num_employees):
    for i in range(num_days):
        model.Add(sum(work[e,d,s] for s in range(2,num_shifts) for d in range(i-6,i))<7)

# 禁止休-班-休模式
# 当前因为上周期班表没有设为变量，所以d-1无法用AddBoolOr
# 下面这个是变通解决上周最后两天分别是 休，休-班这两种情况
for e in range(num_employees):
    for d in range(num_days-1):
        for s in range(1,num_shifts-2):
            if d == 0:
                if (e,d-1,0) in lastweek_assignments:
                    model.AddBoolOr([work[e,d+1,0].Not(),work[e,d,s].Not()])
                elif (e,d-2,0) in lastweek_assignments:
                    model.AddBoolOr([work[e,d,0].Not()])
            else:
                #model.Add(sum([work[e,d-1,0],work[e,d,s],work[e,d+1,0]])<3)
                model.AddBoolOr([work[e,d-1,0].Not(),work[e,d+1,0].Not(),work[e,d,s].Not()])


# 早班班个数约束。从外部读取
night=[7,8,9,10]
M910=[9,10]
for e in range(num_employees):
    model.Add(sum(work[e,d,s] for d in range(num_days) for s in M910)<(min_M910+1))
#    model.Add(sum(work[e,d,s] for d in range(num_days) for s in M910)>(min_M910-1))
    model.Add(sum(work[e,d,s] for d in range(num_days) for s in night)>=(min_night-1))
    model.Add(sum(work[e,d,s] for d in range(num_days) for s in night)<(min_night+1))
# 班次衔接约束
# (前一个班次，后一个班次)
# M9/M10-M1/M2/M3/M4/VM1/VM2/培训
# M8/VM3-VM1
penalized_transitions=[
    (10,1),(10,2),(10,3),(10,4),(10,5),(10,6),(10,12),
    (9,1),(9,2),(9,3),(9,4),(9,5),(9,6),(9,12),
    (7,5),
    (8,5)
]
for previous_shift,next_shift in penalized_transitions:
    for e in range(num_employees):
        for d in range(num_days):
            # 减少work{}的数量
#            if (e,d,next_shift) in work:
                transition=[
                    work[e,d-1,previous_shift].Not(),work[e,d,next_shift].Not()
                ]
                model.AddBoolOr(transition)
#            else:
#                pass

# 休息天数约束。从外部读取
rest_days=num_days-work_days
for e in range(num_employees):
    model.Add(sum(work[e,d,0] for d in range(num_days)  if d not in holi_days)==rest_days)

# 每日班次个数要求.从外部读取
#dayly_cover_demands=[]
# range(1,num_shifts-2)仅为可变班次，不含培训
for s in range(1,num_shifts-2):
    for d in range(num_days):
        works=[work[e,d,s] for e in range(num_employees)]
        min_demand=dayly_cover_demands[d][s-1]
        model.Add(sum(works)>=min_demand)
print("\nstart solver")
# 规划目标
model.Maximize(
    sum(obj_bool_vars[i]*obj_bool_coeffs[i] for i in range(len(obj_bool_vars)))
)
solver=cp_model.CpSolver()
solver.parameters.max_time_in_seconds=400
''' 
# 原始的printer开始，下方用自己写的class替换
solution_printer=cp_model.ObjectiveSolutionPrinter()
status=solver.SolveWithSolutionCallback(model,solution_printer)

# print
if status==cp_model.OPTIMAL or status==cp_model.FEASIBLE:
    print()
    header='      '
    for d in range(num_days):
        header += str(d) + '  ' 
    print(header)
    for e in range(num_employees):
        schedule=''
        for d in range(num_days):
            for s in range(num_shifts):
                if solver.BooleanValue(work[e,d,s]):
                    schedule += shifts[s] + '  '
        print('worker %i:%s' % (e,schedule))
    print()
    print('Penalties:')
    # excel输出
    # 定义字体样式，颜色为红色
    font_style = Font(color='FF0000') # 红色的十六进制代码是 FF0000
    # 定义填充样式，颜色为黄色
    fill_style = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # 创建一个新的工作簿
    wb = Workbook()
    # 选择活动工作表
    ws = wb.active
    for d in range(num_days):
        ws.cell(row=1, column=2+d).value = d+1
    for e in range(num_employees):
        ws.cell(row=2+e, column=1).value = employees[e]
        for d in range(num_days):
            for s in range(num_shifts):
                if solver.BooleanValue(work[e,d,s]):
                    ws.cell(row=2+e, column=2+d).value = shifts[s]
                    #if requests[e,d,s]:
        

    filename = f"solution_00.xlsx"
    # 保存工作簿
    wb.save(filename)
    print(filename)
    wb.close()    
else:
    print("no solution")
# 原始的printer结束
'''
# 下面尝试写一个输出多个可行解的class，与上面的输出可能冲突
# 定义字体样式，颜色为红色
font_style = Font(color='FF0000') # 红色的十六进制代码是 FF0000
# 定义填充样式，颜色为黄色
fill_style = PatternFill(start_color='F2FD99', end_color='F2FD99', fill_type='solid')
fill_style_blue = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
# 创建对齐方式对象并设置水平和垂直居中
align_center = Alignment(horizontal='center', vertical='center')
# 创建边框样式对象，使用默认的细线和黑色
border_style = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
class PartialSolutionPrinter(cp_model.CpSolverSolutionCallback):
    #Print intermediate solutions.
    def __init__(self,shifts,employees,num_days,limit) -> None:
        cp_model.CpSolverSolutionCallback.__init__(self)
        self._shifts=shifts
        self._employees=employees
        self._num_days=num_days
        self._solution_count=0
        self._solution_limit=limit

    def on_solution_callback(self):
        self._solution_count+=1
        rest_meet=0
        duty_meet=0
        # 这里写输出的逻辑
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择活动工作表
        ws = wb.active
        for d in range(self._num_days):
            ws.cell(row=6, column=2+d).value = d+1
            ws.cell(row=6, column=2+d).alignment=align_center
        for e in range(len(self._employees)):
            ws.cell(row=7+e, column=1).value = self._employees[e]
            ws.cell(row=7+e, column=1).alignment=align_center
            ws.cell(row=7+e, column=1).border=border_style
            for d in range(self._num_days):
                for s in range(len(self._shifts)):
                    #这里的取值写法要注意，目前还是自己乱写的
                    if self.BooleanValue(work[e,d,s]):
#                    if solver.BooleanValue(work[e,d,s]):
                        ws.cell(row=7+e, column=2+d).value = self._shifts[s]
                        ws.cell(row=7+e, column=2+d).alignment=align_center
                        ws.cell(row=7+e, column=2+d).border=border_style
                        if any(r[0]==e and r[1]==d and r[2]==s for r in fixed_assignments):
                            ws.cell(row=7+e, column=2+d).fill=fill_style_blue
                        else:
                            if any(r[0]==e and r[1]==d for r in requests):
                                ws.cell(row=7+e, column=2+d).font=font_style
                            if any(r[0]==e and r[1]==d and r[2]==s for r in requests):
                                ws.cell(row=7+e, column=2+d).fill=fill_style
                                if s==0:
                                    rest_meet+=1
                                else:
                                    duty_meet+=1         
        ws.cell(row=1, column=1).value ="score:"
        ws.cell(row=1, column=3).value =self.ObjectiveValue()
        ws.cell(row=2, column=1).value ="conflicts:"
        ws.cell(row=2, column=3).value =self.NumConflicts()
        ws.cell(row=3, column=1).value ="branches:"
        ws.cell(row=3, column=3).value =self.NumBranches()
        ws.cell(row=4, column=1).value ="rest-meet:"
        ws.cell(row=4, column=3).value =str(rest_meet)+" of"+str(rest_count)
        ws.cell(row=5, column=1).value ="shifts-meet:"
        ws.cell(row=5, column=3).value =str(duty_meet)+" of"+str(duty_count)
        filename = f"solution_{self._solution_count}.xlsx"
        # 保存工作簿
        wb.save(filename)
        print(f"\nsolution_{self._solution_count} found")
        print(f"-got score    :{self.ObjectiveValue()}")
        print(f"-rest meet    :{rest_meet} of {rest_count}")
        print(f"-shifts meet  :{duty_meet} of {duty_count}")

        wb.close()
        if self._solution_count>=self._solution_limit:
            print(f"停止！已找到{self._solution_limit}个方案。")
            self.StopSearch()
    def solutionCount(self):
        return self._solution_count
    
# Display the first five solutions.
solution_limit=5
# 用重写的Printer替换默认的
solution_printer=PartialSolutionPrinter(shifts,employees,num_days,solution_limit)
solver.Solve(model,solution_printer)

# Statistics.
print("\nStatistics")
#print(f" - conflicts       : {solver.num_conflicts}")
#print(f" - branches        : {solver.num_branches}")
#print(f" - wall time       : {solver.wall_time}")
#print(f" - solutions found : {solution_printer.solutionCount()}")
   